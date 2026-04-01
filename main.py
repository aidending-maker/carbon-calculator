# ========================================
# 工程碳足跡計算系統 - 主程式（整合版）
# ========================================

import sys
import os
import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from rapidfuzz import fuzz, process
import json
import datetime

# ========================================
# 設定
# ========================================
if len(sys.argv) > 1:
    XML_檔名 = sys.argv[1]
else:
    XML_檔名 = "預算書.xml"

係數資料庫檔 = "資料庫.xlsx"
API_KEY_檔   = "設定.txt"
時間戳記 = datetime.datetime.now().strftime("%Y%m%d_%H%M")

高信心門檻 = 75
低信心門檻 = 45
人工每小時碳排 = 0.0223
人工每工碳排   = 0.1784

忽略關鍵字 = [
    "管理費", "保險費", "營業稅", "薪資", "簽證費", "補償費",
    "職業安全衛生費", "環境保護措施費", "品質管制費",
    "零星工料", "零星費", "工具損耗", "工具費",
    "雜項費用", "雜費", "作業費", "服務費",
    "稅金", "規費", "利潤", "安衛費", "品管費", "環保費", "損耗",
]

def 應該忽略(名稱):
    for k in 忽略關鍵字:
        if k in str(名稱):
            return True
    return False

def 清理名稱(名稱):
    名稱 = str(名稱).strip()
    for 前綴 in ["產品，", "產品,", "材料費，", "材料費,",
                 "工資，", "工資,", "費用，", "費用,"]:
        if 名稱.startswith(前綴):
            名稱 = 名稱[len(前綴):]
    return 名稱.strip()

def 讀取API金鑰():
    try:
        with open(API_KEY_檔, "r", encoding="utf-8") as f:
            for line in f:
                if "ANTHROPIC_API_KEY=" in line:
                    return line.strip().split("=", 1)[1]
    except:
        return None
    return None

# ========================================
# 檢查檔案
# ========================================
print("=" * 50)
print("  工程碳足跡計算系統")
print("=" * 50)

if not os.path.exists(XML_檔名):
    print(f"\n❌ 找不到 XML 檔案：{XML_檔名}")
    sys.exit(1)

if not os.path.exists(係數資料庫檔):
    print(f"\n❌ 找不到係數資料庫：{係數資料庫檔}")
    sys.exit(1)

# ========================================
# 模組一：讀取 XML
# ========================================
print(f"\n📂 模組一：讀取 XML 預算書")

NS = {"ns": "http://pcstd.pcc.gov.tw/2003/eTender"}
tree = ET.parse(XML_檔名)
root = tree.getroot()

標案名稱 = root.find(".//ns:ContractTitle[@language='zh-TW']", NS).text.strip()
機關名稱 = root.find(".//ns:ProcuringEntity[@language='zh-TW']", NS).text.strip()
地點_el  = root.find(".//ns:ContractLocation", NS)
地點     = 地點_el.text.strip() if 地點_el is not None and 地點_el.text else "未填寫"

print(f"  標案：{標案名稱}")
print(f"  機關：{機關名稱}")

safe名稱  = 標案名稱[:20].replace("/","_").replace("\\","_")
輸出前綴  = f"{safe名稱}_{時間戳記}"
清冊檔   = f"{輸出前綴}_資源清冊.xlsx"
比對檔   = f"{輸出前綴}_係數比對.xlsx"
計算檔   = f"{輸出前綴}_碳足跡計算.xlsx"
儀表板檔 = f"{輸出前綴}_儀表板.html"

工項清單 = []

def 解析工項(pay_item, 上層工程="", 層級=0):
    名稱_el = pay_item.find("ns:Description[@language='zh-TW']", NS)
    名稱    = 名稱_el.text.strip() if 名稱_el is not None and 名稱_el.text else ""
    單位_el = pay_item.find("ns:Unit[@language='zh-TW']", NS)
    單位    = 單位_el.text.strip() if 單位_el is not None and 單位_el.text else ""
    數量_el = pay_item.find("ns:Quantity", NS)
    數量    = float(數量_el.text) if 數量_el is not None else 0
    項次    = pay_item.get("itemNo", "")
    種類    = pay_item.get("itemKind", "")
    代碼    = pay_item.get("refItemCode", "").strip()
    單價_el = pay_item.find("ns:Price", NS)
    單價    = float(單價_el.text) if 單價_el is not None and 單價_el.text else 0

    def 取比例(標籤):
        el = pay_item.find(f"ns:{標籤}", NS)
        return float(el.text) if el is not None and el.text else 0

    if 種類 == "mainItem":
        for 子項 in pay_item.findall("ns:PayItem", NS):
            解析工項(子項, 名稱, 層級+1)
    else:
        工項清單.append({
            "所屬工程": 上層工程, "項次": 項次,
            "工項名稱": 名稱, "參考代碼": 代碼,
            "單位": 單位, "數量": 數量, "單價": 單價,
            "種類": 種類,
            "人工比例": 取比例("LabourRatio"),
            "機具比例": 取比例("EquipmentRatio"),
            "材料比例": 取比例("MaterialRatio"),
            "雜項比例": 取比例("MiscellaneaRatio"),
        })
        for 子項 in pay_item.findall("ns:PayItem", NS):
            解析工項(子項, 上層工程, 層級+1)

detail_list = root.find(".//ns:DetailList", NS)
if detail_list is not None:
    for pay_item in detail_list.findall("ns:PayItem", NS):
        解析工項(pay_item)

資源清單 = []

def 判斷資源類型(代碼):
    if 代碼.startswith("L"): return "人工"
    elif 代碼.startswith("E"): return "機具"
    elif 代碼.startswith("M"): return "材料"
    else: return "其他"

for work_item in root.findall(".//ns:WorkItem[@refItemNo]", NS):
    工項參考號 = work_item.get("refItemNo", "")
    工項名稱_el = work_item.find("ns:Description[@language='zh-TW']", NS)
    工項名稱 = 工項名稱_el.text.strip() if 工項名稱_el is not None else ""
    for 資源 in work_item.findall("ns:WorkItem", NS):
        資源代碼 = 資源.get("itemCode", "")
        名稱_el  = 資源.find("ns:Description[@language='zh-TW']", NS)
        名稱     = 名稱_el.text.strip() if 名稱_el is not None else ""
        單位_el  = 資源.find("ns:Unit[@language='zh-TW']", NS)
        單位     = 單位_el.text.strip() if 單位_el is not None and 單位_el.text else ""
        數量_el  = 資源.find("ns:Quantity", NS)
        數量     = float(數量_el.text) if 數量_el is not None else 0
        單價_el  = 資源.find("ns:Price", NS)
        單價     = float(單價_el.text) if 單價_el is not None and 單價_el.text else 0
        備註_el  = 資源.find("ns:Remark", NS)
        備註     = 備註_el.text.strip() if 備註_el is not None and 備註_el.text else ""
        資源清單.append({
            "所屬工項參考號": 工項參考號,
            "所屬工項名稱":   工項名稱,
            "資源代碼":       資源代碼,
            "資源類型":       判斷資源類型(資源代碼),
            "資源名稱":       名稱,
            "單位":           單位,
            "單位數量":       數量,
            "單價":           單價,
            "備註":           備註,
        })

print(f"  ✅ 工項：{len(工項清單)} 個，人機料：{len(資源清單)} 筆")

wb1 = openpyxl.Workbook()
ws_info = wb1.active
ws_info.title = "標案資訊"
for 項, 值 in [("項目","內容"),("標案名稱",標案名稱),("機關名稱",機關名稱),
               ("工程地點",地點),("工項總數",len(工項清單)),("資源總筆數",len(資源清單))]:
    ws_info.append([項, 值])

ws_工項 = wb1.create_sheet("工項清單")
欄位_工項 = ["所屬工程","項次","工項名稱","參考代碼","單位","數量","單價","種類",
             "人工比例","機具比例","材料比例","雜項比例"]
ws_工項.append(欄位_工項)
for 工項 in 工項清單:
    ws_工項.append([工項.get(k,"") for k in 欄位_工項])

ws_資源 = wb1.create_sheet("人機料清單")
欄位_資源 = ["所屬工項參考號","所屬工項名稱","資源代碼","資源類型",
             "資源名稱","單位","單位數量","單價","備註"]
ws_資源.append(欄位_資源)
for 資源 in 資源清單:
    ws_資源.append([資源.get(k,"") for k in 欄位_資源])
    類型 = 資源["資源類型"]
    顏色 = "FFF2CC" if 類型=="人工" else "E2EFDA" if 類型=="機具" else "FCE4D6"
    for cell in ws_資源[ws_資源.max_row]:
        cell.fill = PatternFill("solid", fgColor=顏色)

wb1.save(清冊檔)
print(f"  💾 資源清冊已儲存")

# ========================================
# 模組二：係數比對
# ========================================
print(f"\n🔍 模組二：比對碳係數")

係數清單 = []

def 讀取係數表(sheet, 名稱col, 係數col, 單位col, 來源, header=2, 優先序=99):
    try:
        df = pd.read_excel(係數資料庫檔, sheet_name=sheet, header=header)
        整理 = pd.DataFrame({
            "名稱": df.iloc[:, 名稱col].astype(str).str.strip(),
            "係數": pd.to_numeric(df.iloc[:, 係數col], errors="coerce"),
            "單位": df.iloc[:, 單位col].astype(str).str.strip(),
            "來源": 來源,
            "優先序": 優先序,
        })
        整理 = 整理.dropna(subset=["係數"])
        整理 = 整理[整理["名稱"].str.len() > 1]
        係數清單.append(整理)
        print(f"  ✅ {來源}：{len(整理)} 筆")
    except Exception as e:
        print(f"  ⚠️ {來源} 讀取失敗：{e}")

讀取係數表("資料合併", 1, 2, 3, "環境部", header=1, 優先序=1)
讀取係數表("表五", 2, 5, 4, "工程會", header=2, 優先序=2)
讀取係數表("表四", 2, 4, 5, "水利署", header=2, 優先序=3)
讀取係數表("表七", 14, 20, 21, "機具碳排", header=2, 優先序=4)
讀取係數表("表八", 1, 2, 3, "野溪工程", header=1, 優先序=5)
讀取係數表("表三", 1, 4, 3, "LCBA", header=2, 優先序=6)

全部係數 = pd.concat(係數清單, ignore_index=True)
全部係數 = 全部係數.sort_values("優先序")
全部係數 = 全部係數[全部係數["名稱"].str.len() > 1]
係數名稱清單 = 全部係數["名稱"].tolist()
print(f"  📊 係數資料庫總計：{len(全部係數)} 筆")

def 取得最佳係數(比對名稱):
    符合 = 全部係數[全部係數["名稱"] == 比對名稱]
    if len(符合) == 0: return None
    return 符合.sort_values("優先序").iloc[0]

def AI比對(資源名稱, 候選清單):
    try:
        import anthropic
        api_key = 讀取API金鑰()
        if not api_key: return None, "未匹配"
        client = anthropic.Anthropic(api_key=api_key)
        候選文字 = "\n".join([f"{i+1}. {c}" for i, c in enumerate(候選清單[:10])])
        prompt = f"""你是工程碳足跡專家。請從以下候選係數中，選出最適合「{資源名稱}」的選項。
候選清單：
{候選文字}
請只回答一個數字（1-{min(10,len(候選清單))}）。如果都不適合請回答0。"""
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=10,
            messages=[{"role": "user", "content": prompt}]
        )
        回答 = message.content[0].text.strip()
        選擇 = int(回答) if 回答.isdigit() else 0
        if 1 <= 選擇 <= len(候選清單):
            return 候選清單[選擇-1], "AI比對"
        return None, "未匹配"
    except:
        return None, "未匹配"

比對結果清單 = []
已比對數 = AI比對數 = 未匹配數 = 不需比對數 = 0

for _, row in pd.DataFrame(資源清單).iterrows():
    資源名稱 = str(row.get("資源名稱","")).strip()
    if not 資源名稱 or 資源名稱 == "nan": continue
    資源類型 = str(row.get("資源類型","")).strip()

    if 資源類型 == "人工":
        比對結果清單.append({**row.to_dict(),
            "比對名稱":"人工固定係數","碳係數":0.1784,
            "係數單位":"kgCO₂e/工","係數來源":"固定公式",
            "比對方式":"人工固定","相似度":"－","狀態":"✅ 已匹配"})
        已比對數 += 1
        continue

    if 應該忽略(資源名稱):
        比對結果清單.append({**row.to_dict(),
            "比對名稱":"－","碳係數":"－","係數單位":"－",
            "係數來源":"－","比對方式":"－","相似度":"－","狀態":"⬜ 不需比對"})
        不需比對數 += 1
        continue

    比對 = process.extractOne(清理名稱(資源名稱), 係數名稱清單,
                              scorer=fuzz.token_sort_ratio)

    if 比對 and 比對[1] >= 高信心門檻:
        係數資料 = 取得最佳係數(比對[0])
        比對結果清單.append({**row.to_dict(),
            "比對名稱": 係數資料["名稱"],
            "碳係數":   係數資料["係數"],
            "係數單位": 係數資料["單位"],
            "係數來源": 係數資料["來源"],
            "比對方式": "模糊比對",
            "相似度":   round(比對[1],1),
            "狀態":     "✅ 已匹配"})
        已比對數 += 1
    elif 比對 and 比對[1] >= 低信心門檻:
        候選 = process.extract(清理名稱(資源名稱), 係數名稱清單,
                               scorer=fuzz.token_sort_ratio, limit=5)
        候選名稱 = [c[0] for c in 候選]
        AI選擇, _ = AI比對(資源名稱, 候選名稱)
        if AI選擇:
            係數資料 = 取得最佳係數(AI選擇)
            比對結果清單.append({**row.to_dict(),
                "比對名稱": 係數資料["名稱"],
                "碳係數":   係數資料["係數"],
                "係數單位": 係數資料["單位"],
                "係數來源": 係數資料["來源"],
                "比對方式": "AI比對",
                "相似度":   round(比對[1],1),
                "狀態":     "🤖 AI比對"})
            AI比對數 += 1
        else:
            比對結果清單.append({**row.to_dict(),
                "比對名稱":"","碳係數":"","係數單位":"",
                "係數來源":"","比對方式":"",
                "相似度":round(比對[1],1),"狀態":"❌ 未匹配"})
            未匹配數 += 1
    else:
        比對結果清單.append({**row.to_dict(),
            "比對名稱":"","碳係數":"","係數單位":"",
            "係數來源":"","比對方式":"",
            "相似度":round(比對[1],1) if 比對 else 0,"狀態":"❌ 未匹配"})
        未匹配數 += 1

wb2 = openpyxl.Workbook()
ws_比對 = wb2.active
ws_比對.title = "係數比對結果"
欄位_比對 = ["所屬工項參考號","所屬工項名稱","資源代碼","資源類型",
             "資源名稱","單位","單位數量","單價",
             "比對名稱","碳係數","係數單位","係數來源","比對方式","相似度","狀態"]
for col, 欄 in enumerate(欄位_比對, 1):
    cell = ws_比對.cell(row=1, column=col, value=欄)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")

for row_data in 比對結果清單:
    ws_比對.append([row_data.get(k,"") for k in 欄位_比對])
    狀態 = str(row_data.get("狀態",""))
    顏色 = "E2EFDA" if "✅" in 狀態 else "FFF2CC" if "🤖" in 狀態 else \
           "D9D9D9" if "⬜" in 狀態 else "FCE4D6"
    for cell in ws_比對[ws_比對.max_row]:
        cell.fill = PatternFill("solid", fgColor=顏色)

ws_審核 = wb2.create_sheet("人工審核表")
審核欄位 = ["所屬工項名稱","資源類型","資源名稱","單位","單價",
            "建議比對名稱","建議係數","建議來源",
            "【人工填入】碳係數","【人工填入】單位","【人工填入】來源","備註"]
ws_審核.append(審核欄位)
for row_data in 比對結果清單:
    if "❌" not in str(row_data.get("狀態","")): continue
    ws_審核.append([
        row_data.get("所屬工項名稱",""), row_data.get("資源類型",""),
        row_data.get("資源名稱",""), row_data.get("單位",""),
        row_data.get("單價",""), row_data.get("比對名稱",""),
        row_data.get("碳係數",""), row_data.get("係數來源",""),
        "", "", "", ""
    ])
    for cell in ws_審核[ws_審核.max_row]:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")

wb2.save(比對檔)
print(f"  ✅ 模糊比對：{已比對數} 筆")
print(f"  🤖 AI比對：{AI比對數} 筆")
print(f"  ❌ 未匹配：{未匹配數} 筆")
print(f"  ⬜ 不需比對：{不需比對數} 筆")
print(f"  💾 比對結果已儲存")

人工填入字典 = {}
try:
    df_審核 = pd.read_excel(比對檔, sheet_name="人工審核表")
    for _, row in df_審核.iterrows():
        資源名稱 = str(row.iloc[2]).strip()
        填入係數 = row.iloc[8]
        填入單位 = str(row.iloc[9]).strip() if str(row.iloc[9]) != "nan" else ""
        填入來源 = str(row.iloc[10]).strip() if str(row.iloc[10]) != "nan" else "人工填入"
        if str(填入係數) not in ["nan","","None"] and 資源名稱:
            try:
                人工填入字典[資源名稱] = {
                    "碳係數": float(填入係數),
                    "係數單位": 填入單位,
                    "係數來源": 填入來源,
                }
            except: pass
except: pass

# ========================================
# 模組三：計算碳足跡
# ========================================
print(f"\n🔢 模組三：計算碳足跡")

係數字典 = {}
for row_data in 比對結果清單:
    名稱 = str(row_data.get("資源名稱","")).strip()
    狀態 = str(row_data.get("狀態",""))
    if "✅" in 狀態 or "🤖" in 狀態:
        try:
            係數字典[名稱] = {
                "碳係數": float(row_data.get("碳係數",0)),
                "係數單位": str(row_data.get("係數單位","")),
                "係數來源": str(row_data.get("係數來源","")),
            }
        except: pass

資源碳足跡清單 = []

for 資源 in 資源清單:
    資源名稱 = str(資源.get("資源名稱","")).strip()
    資源類型 = str(資源.get("資源類型","")).strip()
    單位     = str(資源.get("單位","")).strip()
    單位數量 = float(資源.get("單位數量",0) or 0)
    單價     = float(資源.get("單價",0) or 0)
    工項參考號 = str(資源.get("所屬工項參考號","")).strip()
    工項名稱   = str(資源.get("所屬工項名稱","")).strip()

    碳足跡 = 0.0
    計算方式 = ""
    備註 = ""

    if 應該忽略(資源名稱):
        計算方式 = "忽略"
        備註 = "不計入碳足跡"
    elif 資源類型 == "人工":
        if 單位 in ["工","式"]:
            碳足跡 = 單位數量 * 人工每工碳排
            計算方式 = "人工（每工8小時）"
            備註 = f"{單位數量} × 0.1784"
        elif 單位 == "時":
            碳足跡 = 單位數量 * 人工每小時碳排
            計算方式 = "人工（小時）"
            備註 = f"{單位數量} × 0.0223"
        else:
            計算方式 = "⚠️ 待確認"
            備註 = f"人工單位「{單位}」無法自動計算"
    else:
        if 資源名稱 in 人工填入字典:
            係數資料 = 人工填入字典[資源名稱]
            碳足跡 = 單位數量 * 係數資料["碳係數"]
            計算方式 = f"人工填入（{係數資料['係數來源']}）"
            備註 = f"{單位數量} × {係數資料['碳係數']}"
        elif 資源名稱 in 係數字典:
            係數資料 = 係數字典[資源名稱]
            碳足跡 = 單位數量 * 係數資料["碳係數"]
            計算方式 = f"自動比對（{係數資料['係數來源']}）"
            備註 = f"{單位數量} × {係數資料['碳係數']}"
        else:
            計算方式 = "⚠️ 待確認"
            備註 = "找不到係數"

    資源碳足跡清單.append({
        "工項參考號": 工項參考號,
        "工項名稱":   工項名稱,
        "資源類型":   資源類型,
        "資源名稱":   資源名稱,
        "單位":       單位,
        "單位數量":   單位數量,
        "碳足跡(kgCO₂e)": round(碳足跡, 6),
        "計算方式":   計算方式,
        "備註":       備註,
    })

df_資源碳足跡 = pd.DataFrame(資源碳足跡清單)

工項碳足跡清單 = []
for 工項 in 工項清單:
    工項參考號 = str(工項.get("項次","")).strip()
    工項名稱   = str(工項.get("工項名稱","")).strip()
    所屬工程   = str(工項.get("所屬工程","")).strip()
    工項數量   = float(工項.get("數量",1) or 1)
    該工項 = df_資源碳足跡[df_資源碳足跡["工項參考號"]==工項參考號]
    單位碳足跡 = 該工項["碳足跡(kgCO₂e)"].sum()
    總碳足跡   = 單位碳足跡 * 工項數量
    待確認數   = int(該工項["計算方式"].astype(str).str.contains("⚠️").sum())
    工項碳足跡清單.append({
        "所屬工程": 所屬工程,
        "工項參考號": 工項參考號,
        "工項名稱": 工項名稱,
        "工項數量": 工項數量,
        "單位碳足跡(kgCO₂e)": round(單位碳足跡,4),
        "總碳足跡(kgCO₂e)":   round(總碳足跡,4),
        "總碳足跡(tCO₂e)":    round(總碳足跡/1000,6),
        "待確認項目數": 待確認數,
        "備註": "⚠️ 待確認" if 待確認數>0 else "✅ 完整"
    })

df_工項碳足跡 = pd.DataFrame(工項碳足跡清單)

df_工程碳足跡 = df_工項碳足跡.groupby("所屬工程").agg(
    工程碳足跡_kg=("總碳足跡(kgCO₂e)","sum"),
    工程碳足跡_t=("總碳足跡(tCO₂e)","sum"),
    工項數=("工項參考號","count"),
    待確認數=("待確認項目數","sum")
).reset_index()
df_工程碳足跡.columns = ["工程名稱","工程碳足跡(kgCO₂e)","工程碳足跡(tCO₂e)","工項數","待確認項目數"]

專案總碳足跡_t = df_工程碳足跡["工程碳足跡(tCO₂e)"].sum()
待確認總數 = int(df_資源碳足跡["計算方式"].astype(str).str.contains("⚠️").sum())

wb3 = openpyxl.Workbook()

def 寫標題列(ws, 欄位):
    ws.append(欄位)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

ws_總覽 = wb3.active
ws_總覽.title = "專案總覽"
for 項, 值 in [
    ("項目","內容"),
    ("標案名稱", 標案名稱),
    ("主辦機關", 機關名稱),
    ("報告日期", datetime.datetime.now().strftime("%Y年%m月%d日")),
    ("專案總碳足跡(kgCO₂e)", round(專案總碳足跡_t*1000,2)),
    ("專案總碳足跡(tCO₂e)", round(專案總碳足跡_t,4)),
    ("工程數量", len(df_工程碳足跡)),
    ("工項數量", len(df_工項碳足跡)),
    ("資源總筆數", len(df_資源碳足跡)),
    ("待確認項目", 待確認總數),
]:
    ws_總覽.append([項, 值])

ws_工程 = wb3.create_sheet("工程碳足跡")
寫標題列(ws_工程, ["工程名稱","工程碳足跡(kgCO₂e)","工程碳足跡(tCO₂e)","工項數","待確認項目數"])
for _, row in df_工程碳足跡.iterrows():
    ws_工程.append(list(row))

ws_工項 = wb3.create_sheet("工項碳足跡")
寫標題列(ws_工項, list(df_工項碳足跡.columns))
for _, row in df_工項碳足跡.iterrows():
    ws_工項.append(list(row))

ws_明細 = wb3.create_sheet("資源碳足跡明細")
寫標題列(ws_明細, list(df_資源碳足跡.columns))
for _, row in df_資源碳足跡.iterrows():
    ws_明細.append(list(row))
    計算方式 = str(row["計算方式"])
    顏色 = "D9D9D9" if "忽略" in 計算方式 else \
           "FCE4D6" if "⚠️" in 計算方式 else "E2EFDA"
    for cell in ws_明細[ws_明細.max_row]:
        cell.fill = PatternFill("solid", fgColor=顏色)

wb3.save(計算檔)
print(f"  🌿 專案總碳足跡：{專案總碳足跡_t:.4f} tCO₂e")
print(f"  ⚠️ 待確認：{待確認總數} 筆")
print(f"  💾 計算結果已儲存")

# ========================================
# 模組四：產出儀表板
# ========================================
print(f"\n📊 模組四：產出儀表板")

# 計算覆蓋率（排除項目算已掌握）
總資源數 = len(資源碳足跡清單)
已掌握數 = len([r for r in 資源碳足跡清單
                if "⚠️" not in str(r["計算方式"])])
未掌握數 = 總資源數 - 已掌握數
覆蓋率_項目 = round(已掌握數 / 總資源數 * 100, 1) if 總資源數 > 0 else 0

# 金額覆蓋率
總金額 = sum(
    float(工項.get("數量", 0) or 0) * float(工項.get("單價", 0) or 0)
    for 工項 in 工項清單
)
已掌握工項 = set()
for r in 資源碳足跡清單:
    if "⚠️" not in str(r["計算方式"]):
        已掌握工項.add(r["工項參考號"])
已掌握金額 = sum(
    float(工項.get("數量", 0) or 0) * float(工項.get("單價", 0) or 0)
    for 工項 in 工項清單
    if str(工項.get("項次", "")) in 已掌握工項
)
覆蓋率_金額 = round(已掌握金額 / 總金額 * 100, 1) if 總金額 > 0 else 0

df_工程_有效 = df_工程碳足跡[df_工程碳足跡["工程名稱"] != "【專案總計】"].copy()
df_工程_有效["工程碳足跡(tCO₂e)"] = pd.to_numeric(
    df_工程_有效["工程碳足跡(tCO₂e)"], errors="coerce").fillna(0)
df_工程_有效 = df_工程_有效[df_工程_有效["工程碳足跡(tCO₂e)"] > 0]
df_工程_有效 = df_工程_有效.sort_values("工程碳足跡(tCO₂e)", ascending=False)

df_工項_有效 = df_工項碳足跡[
    pd.to_numeric(df_工項碳足跡["總碳足跡(tCO₂e)"],
    errors="coerce").fillna(0) > 0].copy()
df_工項_有效 = df_工項_有效.sort_values("總碳足跡(tCO₂e)", ascending=False)

df_待確認 = df_資源碳足跡[
    df_資源碳足跡["計算方式"].astype(str).str.contains("⚠️|估計")].copy()

工程資料_json = json.dumps([
    {"名稱": str(row["工程名稱"]),
     "碳排_t": round(float(row["工程碳足跡(tCO₂e)"]),4),
     "碳排_kg": round(float(row["工程碳足跡(kgCO₂e)"]),2),
     "占比": round(float(row["工程碳足跡(tCO₂e)"])/專案總碳足跡_t*100,1)
     if 專案總碳足跡_t > 0 else 0}
    for _, row in df_工程_有效.iterrows()
], ensure_ascii=False)

工項資料_json = json.dumps([
    {"工程": str(row.get("所屬工程",""))[:15],
     "工項": str(row.get("工項名稱",""))[:30],
     "數量": str(row.get("工項數量","")),
     "碳排_kg": round(float(row.get("總碳足跡(kgCO₂e)",0)),2),
     "碳排_t": round(float(row.get("總碳足跡(tCO₂e)",0)),4)}
    for _, row in df_工項_有效.iterrows()
], ensure_ascii=False)

待確認_json = json.dumps([
    {"工項": str(row.get("工項名稱",""))[:25],
     "資源": str(row.get("資源名稱",""))[:30],
     "類型": str(row.get("資源類型","")),
     "備註": str(row.get("備註",""))}
    for _, row in df_待確認.iterrows()
], ensure_ascii=False)

報告日期 = datetime.datetime.now().strftime("%Y年%m月%d日")

# 圓形進度條參數（修正版）
總弧長 = 226.2  # 2 × π × 36
項目弧長 = round(總弧長 * 覆蓋率_項目 / 100, 1)
項目空白 = round(總弧長 - 項目弧長, 1)
金額弧長 = round(總弧長 * 覆蓋率_金額 / 100, 1)
金額空白 = round(總弧長 - 金額弧長, 1)
金額顯示 = f"NT$ {已掌握金額:,.0f} / NT$ {總金額:,.0f}" if 總金額 > 0 else "XML 無單價資料"

# 讀取 Logo 轉 base64
import base64
logo_src = ""
logo路徑 = "logo白底.jpg"
if os.path.exists(logo路徑):
    with open(logo路徑, "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode()
    logo_src = f"data:image/png;base64,{logo_b64}"

html = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>工程碳足跡儀表板 - {標案名稱[:20]}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Microsoft JhengHei',sans-serif;background:#f4f7f4;color:#1a3a2a}}
.hdr{{background:#1B4332;color:white;padding:16px 28px;display:flex;justify-content:space-between;align-items:center}}
.hdr-left{{display:flex;align-items:center;gap:16px}}
.hdr-logo{{height:44px;width:auto}}
.hdr-divider{{width:1px;height:40px;background:rgba(255,255,255,0.25)}}
.hdr-proj{{font-size:11px;opacity:0.7;margin-bottom:3px}}
.hdr h1{{font-size:17px;font-weight:700;color:white}}
.hdr-meta{{font-size:11px;opacity:0.8;line-height:1.9;text-align:right}}
.main{{padding:18px 28px}}
.alert{{background:#FAEEDA;border-left:4px solid #BA7517;border-radius:8px;padding:10px 14px;margin-bottom:16px;font-size:12px;color:#633806}}
.cards{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px}}
.card{{background:white;border-radius:10px;padding:16px 18px;border:0.5px solid #d4e6d4;border-top:4px solid #40916C}}
.card.teal{{border-top-color:#1D9E75}}
.card.amber{{border-top-color:#BA7517}}
.card.red{{border-top-color:#E24B4A}}
.card.ok{{border-top-color:#40916C}}
.card-label{{font-size:11px;color:#52796F;margin-bottom:5px}}
.card-value{{font-size:24px;font-weight:700;color:#1B4332}}
.card-unit{{font-size:11px;color:#74C69D;margin-top:2px}}
.section{{background:white;border-radius:10px;padding:16px 20px;border:0.5px solid #d4e6d4;margin-bottom:16px}}
.section h3{{font-size:13px;font-weight:600;color:#1B4332;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #e8f4e8}}
.cov-grid{{display:grid;grid-template-columns:1fr 1fr;gap:28px}}
.cov-item{{display:flex;align-items:center;gap:16px}}
.bar-wrap{{height:5px;background:#e8f4e8;border-radius:3px;margin-top:4px}}
.bar-g{{height:5px;background:#40916C;border-radius:3px}}
.bar-a{{height:5px;background:#BA7517;border-radius:3px}}
.charts{{display:grid;grid-template-columns:2fr 1fr;gap:16px;margin-bottom:16px}}
.chart-box{{background:white;border-radius:10px;padding:16px 20px;border:0.5px solid #d4e6d4}}
.chart-box h3{{font-size:13px;font-weight:600;color:#1B4332;margin-bottom:12px}}
.chart-container{{position:relative;height:270px}}
.rank-section{{background:white;border-radius:10px;padding:16px 20px;border:0.5px solid #d4e6d4;margin-bottom:16px}}
.rank-section h3{{font-size:13px;font-weight:600;color:#1B4332;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid #e8f4e8}}
.rank-item{{display:flex;align-items:center;padding:7px 0;border-bottom:1px solid #f0f7f0}}
.rn{{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;margin-right:10px;flex-shrink:0}}
.r1{{background:#D4AC0D;color:#7D6608}}.r2{{background:#B2BABB;color:#424949}}.r3{{background:#CA6F1E;color:white}}.rx{{background:#e8f4e8;color:#1B4332}}
.rn-name{{flex:1;font-size:12px}}
.rn-bar-wrap{{width:160px;height:6px;background:#e8f4e8;border-radius:3px;margin:0 12px}}
.rn-bar{{height:6px;background:#40916C;border-radius:3px}}
.rn-val{{font-size:11px;font-weight:600;color:#40916C;min-width:90px;text-align:right}}
.tbl-section{{background:white;border-radius:10px;padding:16px 20px;border:0.5px solid #d4e6d4;margin-bottom:16px}}
.tbl-section h3{{font-size:13px;font-weight:600;color:#1B4332;margin-bottom:12px}}
.tab-btns{{display:flex;gap:8px;margin-bottom:12px}}
.tab-btn{{padding:5px 14px;border-radius:6px;border:0.5px solid #d4e6d4;cursor:pointer;font-size:12px;font-family:inherit;background:white;color:#52796F;transition:all 0.2s}}
.tab-btn.active{{background:#1B4332;color:white;border-color:#1B4332}}
.tab-content{{display:none}}.tab-content.active{{display:block}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
thead tr{{background:#1B4332;color:white}}
th{{padding:9px 10px;text-align:left;font-weight:600}}
td{{padding:8px 10px;border-bottom:1px solid #f0f7f0}}
tr:hover td{{background:#f4f9f4}}
.badge{{display:inline-block;padding:2px 7px;border-radius:9px;font-size:11px;font-weight:600}}
.bl{{background:#E3F2FD;color:#1565C0}}.be{{background:#E8F5E9;color:#2E7D32}}.bm{{background:#FFF3E0;color:#E65100}}
.tbl-scroll{{max-height:360px;overflow-y:auto}}
.footer{{text-align:center;padding:14px;font-size:11px;color:#74C69D;background:white;border-radius:10px;border:0.5px solid #d4e6d4}}
</style>
</head>
<body>

<div class="hdr">
  <div class="hdr-left">
    {'<img src="' + logo_src + '" class="hdr-logo" alt="山椒魚永續工程"/><div class="hdr-divider"></div>' if logo_src else ''}
    <div>
      <div class="hdr-proj">工程碳足跡儀表板</div>
      <h1>{標案名稱}</h1>
    </div>
  </div>
  <div class="hdr-meta">
    <div>主辦機關：{機關名稱}</div>
    <div>工程地點：{地點}</div>
    <div>報告日期：{報告日期}</div>
  </div>
</div>

<div class="main">

  {'<div class="alert">⚠️ 本報告有 '+str(待確認總數)+' 筆資源尚未確認係數（以0計入），實際碳足跡可能偏低。</div>' if 待確認總數>0 else ''}

  <div class="cards">
    <div class="card">
      <div class="card-label">專案總碳足跡</div>
      <div class="card-value">{專案總碳足跡_t:.2f}</div>
      <div class="card-unit">公噸 CO₂e（tCO₂e）</div>
    </div>
    <div class="card teal">
      <div class="card-label">工程總數</div>
      <div class="card-value">{len(df_工程_有效)}</div>
      <div class="card-unit">個工程</div>
    </div>
    <div class="card amber">
      <div class="card-label">工項總數</div>
      <div class="card-value">{len(df_工項碳足跡)}</div>
      <div class="card-unit">個工項</div>
    </div>
    <div class="card {'red' if 待確認總數>0 else 'ok'}">
      <div class="card-label">待確認項目</div>
      <div class="card-value">{待確認總數}</div>
      <div class="card-unit">{'筆需補充' if 待確認總數>0 else '筆（完成）'}</div>
    </div>
  </div>

  <div class="section">
    <h3>計算覆蓋率分析</h3>
    <div class="cov-grid">
      <div class="cov-item">
        <svg width="90" height="90" viewBox="0 0 90 90" style="flex-shrink:0">
          <circle cx="45" cy="45" r="36" fill="none" stroke="#e8f4e8" stroke-width="9"/>
          <circle cx="45" cy="45" r="36" fill="none" stroke="#40916C" stroke-width="9"
            stroke-dasharray="{項目弧長} {項目空白}"
            stroke-dashoffset="56.5" stroke-linecap="round"/>
          <text x="45" y="41" text-anchor="middle" style="font-size:16px;font-weight:700;fill:#1B4332">{覆蓋率_項目}%</text>
          <text x="45" y="55" text-anchor="middle" style="font-size:9px;fill:#52796F">項目覆蓋率</text>
        </svg>
        <div style="flex:1">
          <div style="font-size:12px;color:#52796F;margin-bottom:6px">已掌握項目</div>
          <div style="font-size:20px;font-weight:700;color:#1B4332">{已掌握數} <span style="font-size:13px;color:#52796F;font-weight:400">/ {總資源數} 筆</span></div>
          <div style="margin-top:10px">
            <div style="display:flex;justify-content:space-between;font-size:11px;color:#52796F;margin-bottom:3px"><span>已掌握</span><span>{已掌握數} 筆</span></div>
            <div class="bar-wrap"><div class="bar-g" style="width:{覆蓋率_項目}%"></div></div>
            <div style="display:flex;justify-content:space-between;font-size:11px;color:#52796F;margin-top:5px;margin-bottom:3px"><span>待確認</span><span>{未掌握數} 筆</span></div>
            <div class="bar-wrap"><div class="bar-a" style="width:{100-覆蓋率_項目}%"></div></div>
          </div>
        </div>
      </div>
      <div class="cov-item">
        <svg width="90" height="90" viewBox="0 0 90 90" style="flex-shrink:0">
          <circle cx="45" cy="45" r="36" fill="none" stroke="#e8f4e8" stroke-width="9"/>
          <circle cx="45" cy="45" r="36" fill="none" stroke="#1D9E75" stroke-width="9"
            stroke-dasharray="{金額弧長} {金額空白}"
            stroke-dashoffset="56.5" stroke-linecap="round"/>
          <text x="45" y="41" text-anchor="middle" style="font-size:16px;font-weight:700;fill:#1B4332">{覆蓋率_金額}%</text>
          <text x="45" y="55" text-anchor="middle" style="font-size:9px;fill:#52796F">金額覆蓋率</text>
        </svg>
        <div style="flex:1">
          <div style="font-size:12px;color:#52796F;margin-bottom:6px">已掌握金額</div>
          <div style="font-size:20px;font-weight:700;color:#1D9E75">{覆蓋率_金額}%</div>
          <div style="font-size:11px;color:#52796F;margin-top:8px;line-height:1.8">{金額顯示}</div>
        </div>
      </div>
    </div>
  </div>

  <div class="charts">
    <div class="chart-box">
      <h3>各工程碳排放量（tCO₂e）</h3>
      <div class="chart-container"><canvas id="barChart"></canvas></div>
    </div>
    <div class="chart-box">
      <h3>碳排放比例分布</h3>
      <div class="chart-container"><canvas id="pieChart"></canvas></div>
    </div>
  </div>

  <div class="rank-section">
    <h3>碳排放熱點排名</h3>
    <div id="rankList"></div>
  </div>

  <div class="tbl-section">
    <h3>詳細資料</h3>
    <div class="tab-btns">
      <button class="tab-btn active" onclick="switchTab('工項',this)">工項碳足跡</button>
      <button class="tab-btn" onclick="switchTab('待確認',this)">待確認項目{'<span style="background:#E24B4A;color:white;border-radius:9px;padding:1px 6px;font-size:10px;margin-left:4px">'+str(待確認總數)+'</span>' if 待確認總數>0 else ''}</button>
    </div>
    <div id="tab-工項" class="tab-content active">
      <div class="tbl-scroll">
        <table><thead><tr><th>所屬工程</th><th>工項名稱</th><th>數量</th><th>碳足跡(kgCO₂e)</th><th>碳足跡(tCO₂e)</th><th>占比</th></tr></thead>
        <tbody id="工項tbody"></tbody></table>
      </div>
    </div>
    <div id="tab-待確認" class="tab-content">
      <div class="tbl-scroll">
        <table><thead><tr><th>所屬工項</th><th>資源名稱</th><th>類型</th><th>備註</th></tr></thead>
        <tbody id="待確認tbody"></tbody></table>
      </div>
    </div>
  </div>

  <div class="footer">
    山椒魚永續工程股份有限公司 ｜ Formosanus Engineering Sustainable Solutions ｜ 本報告由系統自動產生，數據僅供內部參考 ｜ {報告日期}
  </div>

</div>

<script>
const 工程資料={工程資料_json};
const 工項資料={工項資料_json};
const 待確認資料={待確認_json};
const 總碳排={round(專案總碳足跡_t,4)};
const 顏色=['#1B4332','#40916C','#74C69D','#B7E4C7','#D8F3DC','#52796F','#2D6A4F','#95D5B2','#1D9E75','#081C15'];

new Chart(document.getElementById('barChart').getContext('2d'),{{
  type:'bar',
  data:{{
    labels:工程資料.map(d=>d.名稱.length>14?d.名稱.slice(0,14)+'...':d.名稱),
    datasets:[{{data:工程資料.map(d=>d.碳排_t),backgroundColor:顏色,borderRadius:4}}]
  }},
  options:{{
    responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},
      tooltip:{{callbacks:{{label:ctx=>` ${{ctx.raw}} tCO₂e (${{工程資料[ctx.dataIndex].占比}}%)`}}}}
    }},
    scales:{{
      y:{{beginAtZero:true,title:{{display:true,text:'tCO₂e',color:'#52796F'}},ticks:{{color:'#52796F'}},grid:{{color:'#e8f4e8'}}}},
      x:{{ticks:{{font:{{size:11}},color:'#52796F'}},grid:{{display:false}}}}
    }}
  }}
}});

new Chart(document.getElementById('pieChart').getContext('2d'),{{
  type:'doughnut',
  data:{{
    labels:工程資料.map(d=>d.名稱.length>10?d.名稱.slice(0,10)+'...':d.名稱),
    datasets:[{{data:工程資料.map(d=>d.碳排_t),backgroundColor:顏色,borderWidth:2,borderColor:'#f4f7f4'}}]
  }},
  options:{{
    responsive:true,maintainAspectRatio:false,
    plugins:{{
      legend:{{position:'bottom',labels:{{font:{{size:10}},padding:8,color:'#1a3a2a'}}}},
      tooltip:{{callbacks:{{label:ctx=>` ${{ctx.raw}} tCO₂e (${{工程資料[ctx.dataIndex].占比}}%)`}}}}
    }}
  }}
}});

const rankList=document.getElementById('rankList');
const maxVal=工程資料[0]?.碳排_t||1;
工程資料.forEach((d,i)=>{{
  const pct=(d.碳排_t/maxVal*100).toFixed(0);
  const rc=i===0?'r1':i===1?'r2':i===2?'r3':'rx';
  rankList.innerHTML+=`<div class="rank-item"><div class="rn ${{rc}}">${{i+1}}</div><div class="rn-name">${{d.名稱}}</div><div class="rn-bar-wrap"><div class="rn-bar" style="width:${{pct}}%"></div></div><div class="rn-val">${{d.碳排_t}} tCO₂e &nbsp; ${{d.占比}}%</div></div>`;
}});

const 工項tbody=document.getElementById('工項tbody');
工項資料.forEach(d=>{{
  const 占比=總碳排>0?(d.碳排_t/總碳排*100).toFixed(1):'0.0';
  工項tbody.innerHTML+=`<tr><td>${{d.工程}}</td><td>${{d.工項}}</td><td>${{d.數量}}</td><td>${{d.碳排_kg.toFixed(2)}}</td><td><strong>${{d.碳排_t}}</strong></td><td><div style="display:flex;align-items:center;gap:5px"><div style="width:50px;height:5px;background:#e8f4e8;border-radius:2px"><div style="width:${{Math.min(占比*3,100)}}%;height:5px;background:#40916C;border-radius:2px"></div></div>${{占比}}%</div></td></tr>`;
}});

const 待確認tbody=document.getElementById('待確認tbody');
if(待確認資料.length===0){{
  待確認tbody.innerHTML='<tr><td colspan="4" style="text-align:center;color:#40916C;padding:16px">✅ 所有項目均已確認係數</td></tr>';
}}else{{
  待確認資料.forEach(d=>{{
    const tc=d.類型==='人工'?'bl':d.類型==='機具'?'be':'bm';
    待確認tbody.innerHTML+=`<tr><td>${{d.工項}}</td><td>${{d.資源}}</td><td><span class="badge ${{tc}}">${{d.類型}}</span></td><td style="color:#854F0B">${{d.備註}}</td></tr>`;
  }});
}}

function switchTab(name,btn){{
  document.querySelectorAll('.tab-content').forEach(el=>el.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(el=>el.classList.remove('active'));
  document.getElementById('tab-'+name).classList.add('active');
  btn.classList.add('active');
}}
</script>
</body>
</html>"""

with open(儀表板檔, "w", encoding="utf-8") as f:
    f.write(html)

print(f"  💾 儀表板已儲存")

# ========================================
# 完成總結
# ========================================
print(f"""
{'='*50}
  計算完成！
{'='*50}
  標案：{標案名稱[:30]}
  總碳足跡：{專案總碳足跡_t:.4f} tCO₂e
  項目覆蓋率：{覆蓋率_項目}%（{已掌握數}/{總資源數} 筆）
  金額覆蓋率：{覆蓋率_金額}%
  工程數：{len(df_工程_有效)} 個
  工項數：{len(df_工項碳足跡)} 項
  待確認：{待確認總數} 筆

  產出檔案：
  📋 {清冊檔}
  🔍 {比對檔}
  📊 {計算檔}
  🌐 {儀表板檔}

  請用瀏覽器打開儀表板檔案查看結果！
{'='*50}
""")
